from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Iterable

try:
    import tkinter as tk
    from tkinter import ttk
except Exception:  # noqa: BLE001
    tk = None
    ttk = None

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

TARGET_AUS = [
    "AU01",
    "AU02",
    "AU04",
    "AU05",
    "AU06",
    "AU07",
    "AU09",
    "AU10",
    "AU12",
    "AU14",
    "AU15",
    "AU17",
    "AU18",
    "AU20",
    "AU21",
    "AU23",
    "AU25",
    "AU26",
    "AU45",
]

SERVICE_COLUMNS = ["frame", "face_id", "timestamp", "confidence", "success"]

EMOTION_ALIASES = {
    "anger": ["гнев", "anger"],
    "sadness": ["грусть", "sad", "sadness"],
    "joy": ["радость", "joy", "happy", "happiness"],
    "surprise": ["удивление", "surprise"],
}

EMOTION_AU_MAP = {
    "anger": {"AU02", "AU04", "AU05", "AU07", "AU23"},
    "sadness": {"AU01", "AU04", "AU15"},
    "joy": {"AU06", "AU12"},
    "surprise": {"AU01", "AU02", "AU05", "AU26"},
}

FILL_C_NON_ZERO = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_R_HIGH = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_EMOTION_HEADER = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
BOLD = Font(bold=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Обработка CSV OpenFace в единый Excel-файл")
    parser.add_argument(
        "--input-root",
        type=Path,
        default=Path(__file__).resolve().parent / "processed_openface",
        help="Папка с подпапками actor/respondent",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("openface_processed.xlsx"),
        help="Имя общего Excel-файла",
    )
    parser.add_argument("--mode", choices=("actor", "respondent"), default=None)
    parser.add_argument("--respondent", default=None, help="Имя респондента для режима respondent")
    parser.add_argument("--no-gui", action="store_true", help="Отключить GUI-меню выбора")
    return parser.parse_args()


def prompt_choice_console(title: str, options: list[str]) -> str:
    print(title)
    for idx, option in enumerate(options, start=1):
        print(f"{idx}. {option}")

    while True:
        raw = input("Введите номер: ").strip()
        if raw.isdigit():
            number = int(raw)
            if 1 <= number <= len(options):
                return options[number - 1]
        print("Некорректный выбор, попробуйте снова.")


def prompt_choice_gui(title: str, options: list[str]) -> str | None:
    if tk is None or ttk is None:
        return None

    try:
        root = tk.Tk()
    except Exception:  # noqa: BLE001
        return None

    root.title(title)
    root.resizable(False, False)

    selected = tk.StringVar(value=options[0])
    result: dict[str, str | None] = {"value": None}

    frame = ttk.Frame(root, padding=14)
    frame.grid(row=0, column=0)

    ttk.Label(frame, text=title).grid(row=0, column=0, sticky="w")
    combo = ttk.Combobox(frame, values=options, textvariable=selected, state="readonly", width=40)
    combo.grid(row=1, column=0, pady=(8, 12), sticky="w")

    def on_ok() -> None:
        result["value"] = selected.get()
        root.destroy()

    def on_cancel() -> None:
        result["value"] = None
        root.destroy()

    buttons = ttk.Frame(frame)
    buttons.grid(row=2, column=0, sticky="e")
    ttk.Button(buttons, text="ОК", command=on_ok).grid(row=0, column=0, padx=(0, 8))
    ttk.Button(buttons, text="Отмена", command=on_cancel).grid(row=0, column=1)

    root.protocol("WM_DELETE_WINDOW", on_cancel)
    root.mainloop()
    return result["value"]


def prompt_choice(title: str, options: list[str], use_gui: bool) -> str:
    if use_gui:
        picked = prompt_choice_gui(title, options)
        if picked:
            return picked
    return prompt_choice_console(title, options)


def resolve_mode(args: argparse.Namespace) -> str:
    if args.mode:
        return args.mode
    return prompt_choice("Выберите режим:", ["actor", "respondent"], use_gui=not args.no_gui)


def read_csv_rows(csv_path: Path) -> tuple[list[str], list[dict[str, str]]]:
    for encoding in ("utf-8-sig", "cp1251", "utf-8"):
        try:
            with csv_path.open("r", encoding=encoding, newline="") as fh:
                reader = csv.DictReader(fh)
                if reader.fieldnames is None:
                    raise ValueError("Пустой или повреждённый CSV")
                fieldnames = [name.strip() for name in reader.fieldnames]
                rows: list[dict[str, str]] = []
                for row in reader:
                    cleaned = {str(k).strip(): (v if v is not None else "") for k, v in row.items()}
                    rows.append(cleaned)
                return fieldnames, rows
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("csv", b"", 0, 1, f"Не удалось декодировать {csv_path}")


def detect_emotion(name: str) -> str | None:
    low = name.lower()
    for emotion, aliases in EMOTION_ALIASES.items():
        if any(alias in low for alias in aliases):
            return emotion
    return None


def parse_file_metadata(csv_path: Path, mode: str) -> tuple[str, str]:
    stem_parts = [part for part in re.split(r"[_\-\s]+", csv_path.stem) if part]
    stimulus = stem_parts[0] if stem_parts else csv_path.stem

    if mode == "actor":
        person = csv_path.parent.name
    else:
        parts = list(csv_path.parts)
        person = csv_path.parent.name
        for idx, part in enumerate(parts):
            if part.lower() == "respondent" and idx + 1 < len(parts):
                person = parts[idx + 1]
                break
        if person.lower().startswith("attempt") and csv_path.parent.parent.name:
            person = csv_path.parent.parent.name

    return stimulus, person


def build_sheet_name(stimulus: str, mode: str, person: str) -> str:
    full = f"{stimulus}_{mode}_{person}"
    invalid = set('[]:*?/\\')
    sanitized = "".join("_" if ch in invalid else ch for ch in full)
    return sanitized[:31]


def collect_au_columns(columns: Iterable[str]) -> list[str]:
    present = set(columns)
    selected: list[str] = []
    for au in TARGET_AUS:
        r_col = f"{au}_r"
        c_col = f"{au}_c"
        if r_col in present:
            selected.append(r_col)
        if c_col in present:
            selected.append(c_col)
    return selected



def build_candidate_roots(root: Path) -> list[Path]:
    project_root = Path(__file__).resolve().parent.parent
    candidates = [
        root,
        Path.cwd() / root,
        project_root / root,
        Path("processed_openface"),
        Path("processed openface"),
        Path("data"),
        project_root / "processed_openface",
        project_root / "processed openface",
        project_root / "data",
    ]

    unique: list[Path] = []
    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate.resolve(strict=False))
        if key in seen:
            continue
        seen.add(key)
        unique.append(candidate)
    return unique

def select_csv_files(
    root: Path,
    mode: str,
    respondent_name: str | None,
    use_gui: bool,
) -> tuple[list[Path], str | None]:
    candidate_roots = build_candidate_roots(root)

    mode_root = None
    for base_root in candidate_roots:
        if base_root.is_dir() and base_root.name.lower() == mode:
            mode_root = base_root
            break

        possible = base_root / mode
        if possible.is_dir():
            mode_root = possible
            break

    if mode_root is None:
        checked = ", ".join(str((base / mode)) for base in candidate_roots)
        raise FileNotFoundError(f"Папка не найдена для режима '{mode}'. Проверены пути: {checked}")

    if mode == "actor":
        return sorted(mode_root.rglob("*.csv")), None

    respondent_dirs = sorted([p for p in mode_root.iterdir() if p.is_dir()])
    respondent_files = sorted(mode_root.glob("*.csv"))

    if respondent_name:
        target_dir = mode_root / respondent_name
        if target_dir.is_dir():
            return sorted(target_dir.rglob("*.csv")), respondent_name
        matched = [p for p in respondent_files if respondent_name.lower() in p.stem.lower()]
        return matched, respondent_name

    options = [d.name for d in respondent_dirs]
    if options:
        picked = prompt_choice("Выберите респондента:", options, use_gui=use_gui)
        return sorted((mode_root / picked).rglob("*.csv")), picked

    names = sorted({p.stem.split("_")[0] for p in respondent_files})
    if not names:
        return [], None
    picked = prompt_choice("Выберите респондента:", names, use_gui=use_gui)
    selected = [p for p in respondent_files if p.stem.lower().startswith(picked.lower())]
    return selected, picked


def print_progress(current: int, total: int) -> None:
    ratio = current / total if total else 1
    percent = int(ratio * 100)
    width = 32
    filled = int(width * ratio)
    bar = "█" * filled + "-" * (width - filled)
    print(f"\rПрогресс обработки: [{bar}] {percent}% ({current}/{total})", end="", flush=True)
    if current == total:
        print()


def prepare_workbook(path: Path):
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
    return wb


def autosize_columns(ws, max_col: int, max_row: int) -> None:
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)


def write_sheet(
    wb,
    csv_path: Path,
    mode: str,
    selected_columns: list[str],
    rows: list[dict[str, str]],
    respondent_name: str | None = None,
) -> None:
    stimulus, person = parse_file_metadata(csv_path, mode)
    if mode == "respondent" and respondent_name:
        person = respondent_name
    sheet_name = build_sheet_name(stimulus, mode, person)

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)

    header_row = 1
    for col_idx, column in enumerate(selected_columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=column)
        cell.font = BOLD

    emotion = detect_emotion(csv_path.stem)
    emotion_aus = EMOTION_AU_MAP.get(emotion, set()) if emotion else set()

    for row_offset, row in enumerate(rows, start=1):
        excel_row = header_row + row_offset
        for col_idx, column in enumerate(selected_columns, start=1):
            value_raw = row.get(column, "")
            ws.cell(row=excel_row, column=col_idx, value=value_raw)

            if column.endswith("_c"):
                try:
                    if float(value_raw) != 0:
                        ws.cell(row=excel_row, column=col_idx).fill = FILL_C_NON_ZERO
                except ValueError:
                    continue
            if column.endswith("_r"):
                try:
                    if float(value_raw) > 0.6:
                        ws.cell(row=excel_row, column=col_idx).fill = FILL_R_HIGH
                except ValueError:
                    continue

    for col_idx, column in enumerate(selected_columns, start=1):
        if re.match(r"^AU\d{2}_[rc]$", column):
            au = column[:4]
            if au in emotion_aus:
                ws.cell(row=header_row, column=col_idx).fill = FILL_EMOTION_HEADER

    last_row = header_row + len(rows)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(selected_columns))}{last_row}"
    autosize_columns(ws, len(selected_columns), last_row)


def main() -> None:
    args = parse_args()
    mode = resolve_mode(args)

    csv_files, respondent_name = select_csv_files(
        args.input_root,
        mode,
        args.respondent,
        use_gui=not args.no_gui,
    )
    if not csv_files:
        print("[WARN] Нет CSV-файлов для обработки.")
        return

    wb = prepare_workbook(args.output)

    processed = 0
    total = len(csv_files)
    done = 0
    for csv_path in csv_files:
        try:
            columns, rows = read_csv_rows(csv_path)
            selected = [col for col in SERVICE_COLUMNS if col in columns] + collect_au_columns(columns)
            if not selected:
                print(f"[WARN] Пропуск файла без нужных колонок: {csv_path}")
                continue
            write_sheet(wb, csv_path, mode, selected, rows, respondent_name=respondent_name)
            processed += 1
            print(f"[OK] {csv_path}")
        except Exception as exc:  # noqa: BLE001
            print(f"[WARN] Пропуск повреждённого/неподходящего файла {csv_path}: {exc}")
        finally:
            done += 1
            print_progress(done, total)

    wb.save(args.output)
    print(f"Готово. Обработано листов: {processed}. Файл: {args.output}")
    if mode == "respondent" and respondent_name:
        print(f"Выбран респондент: {respondent_name}")


if __name__ == "__main__":
    main()

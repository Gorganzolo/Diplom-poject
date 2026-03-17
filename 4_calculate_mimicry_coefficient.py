from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
import math

try:
    import tkinter as tk
    from tkinter import ttk, messagebox
except Exception:  # noqa: BLE001
    tk = None
    ttk = None
    messagebox = None

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


INPUT_FILE = Path("openface_processed.xlsx")
OUTPUT_FILE = Path("mimicry_results.xlsx")
RESULT_SHEET = "Результаты"
ALL_RESPONDENTS_OPTION = "[Все респонденты]"

# Цвета
FILL_HEADER = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")

# Базовый фон строк
FILL_ROW_LIGHT = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
FILL_ROW_DARK = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")

# AU-значения
FILL_AU_LIGHT = PatternFill(fill_type="solid", start_color="F4CCCC", end_color="F4CCCC")
FILL_AU_DARK = PatternFill(fill_type="solid", start_color="EA9999", end_color="EA9999")

# Суммы
FILL_SUM_LIGHT = PatternFill(fill_type="solid", start_color="D9EAD3", end_color="D9EAD3")
FILL_SUM_DARK = PatternFill(fill_type="solid", start_color="B6D7A8", end_color="B6D7A8")

# Средние
FILL_MEAN_LIGHT = PatternFill(fill_type="solid", start_color="CFE2F3", end_color="CFE2F3")
FILL_MEAN_DARK = PatternFill(fill_type="solid", start_color="9FC5E8", end_color="9FC5E8")
BOLD = Font(bold=True)

# Полный набор AU для вывода в таблицу
ALL_AUS = [
    "AU01_r",
    "AU02_r",
    "AU04_r",
    "AU05_r",
    "AU06_r",
    "AU07_r",
    "AU12_r",
    "AU15_r",
    "AU23_r",
    "AU26_r",
]

EMOTIONS: dict[str, dict[str, Any]] = {
    "anger": {
        "title": "гнев",
        "actor_sheet": "гнев_actor_actor",
        "aus": ["AU02_r", "AU04_r", "AU05_r", "AU07_r", "AU23_r"],
        "anchor_frames": [919, 1493],
        "respondent_range": (115, 1812),
    },
    "sadness": {
        "title": "грусть",
        "actor_sheet": "грусть_actor_actor",
        "aus": ["AU01_r", "AU04_r", "AU15_r"],
        "anchor_frames": [32, 277],
        "respondent_range": (1996, 3744),
    },
    "joy": {
        "title": "радость",
        "actor_sheet": "радость_actor_actor",
        "aus": ["AU06_r", "AU12_r"],  # считаем по этим AU
        "anchor_frames": [482, 630],
        "respondent_range": (3886, 5126),
    },
    "surprise": {
        "title": "удивление",
        "actor_sheet": "удивление_actor_actor",
        "aus": ["AU01_r", "AU02_r", "AU05_r", "AU26_r"],
        "anchor_frames": [1829, 1271],
        "respondent_range": (5269, 7500),
    },
}

# Для вывода в таблицу joy тоже кладём значения в AU12_r, AU06_r не выводим в твоём шаблоне не было.
# Если нужен и AU06_r в таблице — добавлю отдельной колонкой.
DISPLAY_AUS = [
    "AU01_r",
    "AU02_r",
    "AU04_r",
    "AU05_r",
    "AU06_r",
    "AU07_r",
    "AU12_r",
    "AU15_r",
    "AU23_r",
    "AU26_r",
]


@dataclass
class SegmentResult:
    frame: int | None
    values: dict[str, float]
    total: float
    mean: float


def normalize_header(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def to_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def build_header_map(ws) -> tuple[dict[str, int], int]:
    for row_idx in range(1, min(20, ws.max_row) + 1):
        values = [normalize_header(ws.cell(row=row_idx, column=col).value) for col in range(1, ws.max_column + 1)]
        lowered = {v.lower() for v in values if v}
        if "frame" in lowered:
            header_map = {}
            for col_idx, value in enumerate(values, start=1):
                if value:
                    header_map[value] = col_idx
            return header_map, row_idx
    raise ValueError(f"Не удалось найти строку заголовков на листе '{ws.title}'")


def find_row_by_frame(ws, header_map: dict[str, int], data_start_row: int, target_frame: int) -> int | None:
    frame_col = header_map.get("frame")
    if frame_col is None:
        return None

    for row_idx in range(data_start_row, ws.max_row + 1):
        frame_value = ws.cell(row=row_idx, column=frame_col).value
        try:
            if int(float(frame_value)) == int(target_frame):
                return row_idx
        except (TypeError, ValueError):
            continue
    return None


def extract_au_values(ws, row_idx: int, header_map: dict[str, int], aus: list[str]) -> dict[str, float]:
    values: dict[str, float] = {}
    for au in aus:
        col = header_map.get(au)
        values[au] = to_float(ws.cell(row=row_idx, column=col).value) if col else 0.0
    return values


def compute_segment(values: dict[str, float]) -> tuple[float, float]:
    nums = list(values.values())
    if not nums:
        return 0.0, 0.0
    total = sum(nums)
    mean = total / len(nums)
    return total, mean


def get_actor_result(ws, emotion_key: str, anchor_frame: int) -> SegmentResult:
    info = EMOTIONS[emotion_key]
    aus = info["aus"]

    header_map, header_row = build_header_map(ws)
    data_start_row = header_row + 1

    row_idx = find_row_by_frame(ws, header_map, data_start_row, anchor_frame)
    if row_idx is None:
        raise ValueError(f"На листе '{ws.title}' не найден frame={anchor_frame}")

    values = extract_au_values(ws, row_idx, header_map, aus)
    total, mean = compute_segment(values)
    return SegmentResult(frame=anchor_frame, values=values, total=total, mean=mean)


def get_respondent_result(ws, emotion_key: str, anchor_frame: int) -> SegmentResult:
    info = EMOTIONS[emotion_key]
    aus = info["aus"]
    range_start, range_end = info["respondent_range"]

    header_map, _ = build_header_map(ws)

    # Логика по твоему правилу:
    # старт эмоции у респондента + опорный кадр актёра
    search_start = range_start + anchor_frame
    search_end = min(search_start + 150, range_end, ws.max_row)

    if search_start > ws.max_row:
        raise ValueError(
            f"Старт поиска выходит за пределы листа: start={search_start}, emotion={emotion_key}"
        )

    if search_start > search_end:
        raise ValueError(
            f"Пустой диапазон поиска у респондента: start={search_start}, end={search_end}, emotion={emotion_key}"
        )

    best_row: int | None = None
    best_values: dict[str, float] = {}
    best_total = -math.inf
    best_mean = -math.inf

    frame_col = header_map.get("frame")

    for row_idx in range(search_start, search_end + 1):
        values = extract_au_values(ws, row_idx, header_map, aus)
        total, mean = compute_segment(values)
        if mean > best_mean:
            best_row = row_idx
            best_values = values
            best_total = total
            best_mean = mean

    if best_row is None:
        raise ValueError(f"Не найден максимум у респондента для emotion={emotion_key}, anchor={anchor_frame}")

    frame_value = ws.cell(row=best_row, column=frame_col).value if frame_col else None
    try:
        best_frame = int(float(frame_value))
    except (TypeError, ValueError):
        best_frame = None

    return SegmentResult(
        frame=best_frame,
        values=best_values,
        total=best_total,
        mean=best_mean,
    )


def get_respondent_sheets(sheet_names: list[str]) -> list[str]:
    return [name for name in sheet_names if name.startswith("face_respondent_")]


def extract_respondent_name(sheet_name: str) -> str:
    prefix = "face_respondent_"
    return sheet_name[len(prefix):] if sheet_name.startswith(prefix) else sheet_name


def choose_respondents_gui(options: list[str]) -> list[str] | None:
    if tk is None or ttk is None:
        return None

    result = {"value": None}

    root = tk.Tk()
    root.title("Выбор респондента")
    root.resizable(False, False)

    frame = ttk.Frame(root, padding=14)
    frame.grid(row=0, column=0)

    ttk.Label(frame, text="Выберите респондента:").grid(row=0, column=0, sticky="w")

    selected = tk.StringVar(value=options[0])
    combo = ttk.Combobox(frame, textvariable=selected, values=options, state="readonly", width=40)
    combo.grid(row=1, column=0, pady=(8, 12), sticky="we")

    def on_ok() -> None:
        result["value"] = selected.get()
        root.destroy()

    def on_cancel() -> None:
        result["value"] = None
        root.destroy()

    buttons = ttk.Frame(frame)
    buttons.grid(row=2, column=0, sticky="e")

    ttk.Button(buttons, text="ОК", command=on_ok).grid(row=0, column=0, padx=(0, 8))
    ttk.Button(buttons, text="Отмена", command=on_cancel).grid(row=0, column=1)

    root.protocol("WM_DELETE_WINDOW", on_cancel)
    root.mainloop()

    if result["value"] is None:
        return None
    if result["value"] == ALL_RESPONDENTS_OPTION:
        return [opt for opt in options if opt != ALL_RESPONDENTS_OPTION]
    return [result["value"]]


def format_display_value(value: float | None) -> Any:
    if value is None:
        return "---"
    return value

def get_block_index(data_row_idx: int) -> int:
    """
    data_row_idx — номер строки с данными, начиная со 2 строки листа.
    Каждые 11 строк = один блок.
    """
    return (data_row_idx - 2) // 11


def is_dark_block(data_row_idx: int) -> bool:
    return get_block_index(data_row_idx) % 2 == 1


def get_row_fill(data_row_idx: int) -> PatternFill:
    return FILL_ROW_DARK if is_dark_block(data_row_idx) else FILL_ROW_LIGHT


def get_au_fill(data_row_idx: int) -> PatternFill:
    return FILL_AU_DARK if is_dark_block(data_row_idx) else FILL_AU_LIGHT


def get_sum_fill(data_row_idx: int) -> PatternFill:
    return FILL_SUM_DARK if is_dark_block(data_row_idx) else FILL_SUM_LIGHT


def get_mean_fill(data_row_idx: int) -> PatternFill:
    return FILL_MEAN_DARK if is_dark_block(data_row_idx) else FILL_MEAN_LIGHT

def create_or_open_output() -> tuple[Any, Any]:
    if OUTPUT_FILE.exists():
        wb = load_workbook(OUTPUT_FILE)
        if RESULT_SHEET in wb.sheetnames:
            ws = wb[RESULT_SHEET]
        else:
            ws = wb.create_sheet(RESULT_SHEET)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = RESULT_SHEET
    return wb, ws


def ensure_headers(ws) -> None:
    headers = [
        "имя респондента",
        "эмоция",
        "опорный кадр актёр",
        "макс кадр респондент",
        "А AU01_r",
        "А AU02_r",
        "А AU04_r",
        "А AU05_r",
        "А AU06_r",
        "А AU07_r",
        "А AU12_r",
        "А AU15_r",
        "А AU23_r",
        "А AU26_r",
        "Сумма актёра",
        "Среднее актёра",
        "Р AU01_r",
        "Р AU02_r",
        "Р AU04_r",
        "Р AU05_r",
        "Р AU06_r",
        "Р AU07_r",
        "Р AU12_r",
        "Р AU15_r",
        "Р AU23_r",
        "Р AU26_r",
        "Сумма респондента",
        "Среднее респондента",
    ]

    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = BOLD
            cell.fill = FILL_HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        return

    existing = [ws.cell(row=1, column=i).value for i in range(1, len(headers) + 1)]
    if existing != headers:
        ws.delete_rows(1, ws.max_row)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = BOLD
            cell.fill = FILL_HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center")


def remove_existing_rows_for_respondent(ws, respondent_name: str) -> None:
    rows_to_delete = []
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        if str(value).strip() == respondent_name:
            rows_to_delete.append(row_idx)

    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, 1)


def build_display_row(
    respondent_name: str,
    emotion_title: str,
    anchor_frame: int,
    actor_result: SegmentResult,
    respondent_result: SegmentResult,
    emotion_aus: list[str],
) -> list[Any]:
    actor_display = []
    respondent_display = []

    for au in DISPLAY_AUS:
        if au in emotion_aus:
            actor_display.append(format_display_value(actor_result.values.get(au, 0.0)))
            respondent_display.append(format_display_value(respondent_result.values.get(au, 0.0)))
        else:
            actor_display.append("---")
            respondent_display.append("---")

    return [
        respondent_name,
        emotion_title,
        anchor_frame,
        respondent_result.frame,
        *actor_display,
        actor_result.total,
        actor_result.mean,
        *respondent_display,
        respondent_result.total,
        respondent_result.mean,
    ]

def style_result_row(ws, row_idx: int) -> None:
    """
    Форматирует одну строку результатов.
    Столбцы:
    1-4   служебные
    5-13  AU актёра
    14    сумма актёра
    15    среднее актёра
    16-24 AU респондента
    25    сумма респондента
    26    среднее респондента
    """
    row_fill = get_row_fill(row_idx)
    au_fill = get_au_fill(row_idx)
    sum_fill = get_sum_fill(row_idx)
    mean_fill = get_mean_fill(row_idx)

    actor_au_cols = range(5, 15)
    respondent_au_cols = range(17, 27)
    sum_cols = [15, 27]
    mean_cols = [16, 28]

    # Сначала красим весь ряд базовым серым
    for col_idx in range(1, 27):
        ws.cell(row=row_idx, column=col_idx).fill = row_fill

    # AU-столбцы: числа красные, "---" остаются серыми
    for col_idx in list(actor_au_cols) + list(respondent_au_cols):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value
        if value != "---" and value is not None and value != "":
            cell.fill = au_fill
        else:
            cell.fill = row_fill

    # Суммы
    for col_idx in sum_cols:
        ws.cell(row=row_idx, column=col_idx).fill = sum_fill

    # Средние
    for col_idx in mean_cols:
        ws.cell(row=row_idx, column=col_idx).fill = mean_fill

def append_rows_for_respondent(ws, source_wb, respondent_sheet_name: str) -> None:
    respondent_name = extract_respondent_name(respondent_sheet_name)
    respondent_ws = source_wb[respondent_sheet_name]

    start_row = ws.max_row + 1

    for emotion_key, info in EMOTIONS.items():
        actor_sheet_name = info["actor_sheet"]
        if actor_sheet_name not in source_wb.sheetnames:
            print(f"[WARN] Не найден лист актёра: {actor_sheet_name}")
            continue

        actor_ws = source_wb[actor_sheet_name]
        emotion_aus = info["aus"]

        for anchor_frame in info["anchor_frames"]:
            try:
                actor_result = get_actor_result(actor_ws, emotion_key, anchor_frame)
                respondent_result = get_respondent_result(respondent_ws, emotion_key, anchor_frame)
                row_data = build_display_row(
                    respondent_name=respondent_name,
                    emotion_title=info["title"],
                    anchor_frame=anchor_frame,
                    actor_result=actor_result,
                    respondent_result=respondent_result,
                    emotion_aus=emotion_aus,
                )

                row_idx = ws.max_row + 1
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                style_result_row(ws, row_idx)

            except Exception as exc:  # noqa: BLE001
                print(f"[WARN] {info['title']} / frame {anchor_frame}: {exc}")


def autosize(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 24)


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Не найден файл: {INPUT_FILE}")

    if tk is None or ttk is None:
        raise RuntimeError("Tkinter недоступен. Нужен GUI для выбора респондентов.")

    source_wb = load_workbook(INPUT_FILE, data_only=True)

    missing_actor_sheets = [info["actor_sheet"] for info in EMOTIONS.values() if info["actor_sheet"] not in source_wb.sheetnames]
    if missing_actor_sheets:
        raise ValueError("Не найдены листы актёра:\n" + "\n".join(missing_actor_sheets))

    respondent_sheets = get_respondent_sheets(source_wb.sheetnames)
    if not respondent_sheets:
        raise ValueError("Не найдены листы респондентов вида 'face_respondent_имя'")

    respondent_options = [ALL_RESPONDENTS_OPTION] + sorted(respondent_sheets)
    selected_respondents = choose_respondents_gui(respondent_options)
    if not selected_respondents:
        return

    out_wb, out_ws = create_or_open_output()
    ensure_headers(out_ws)

    for respondent_sheet in selected_respondents:
        respondent_name = extract_respondent_name(respondent_sheet)
        remove_existing_rows_for_respondent(out_ws, respondent_name)
        append_rows_for_respondent(out_ws, source_wb, respondent_sheet)

    autosize(out_ws)
    out_ws.freeze_panes = "A2"
    out_wb.save(OUTPUT_FILE)

    if messagebox is not None:
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Готово", f"Результат сохранён в файл:\n{OUTPUT_FILE.resolve()}")
        root.destroy()
    else:
        print(f"Готово. Результат сохранён: {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()